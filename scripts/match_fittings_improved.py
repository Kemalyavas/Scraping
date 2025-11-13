"""
IMPROVED Fittings Matcher
Uses merged Heizmann data with better thread extraction
"""

import json
import re
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ImprovedFittingsMatcher:
    """Match fittings with improved data"""
    
    def __init__(self, balflex_file: str, heizmann_file: str):
        self.balflex_fittings = self._load_json(balflex_file)
        self.heizmann_fittings = self._load_json(heizmann_file)
        self.matches = []
        
    def _load_json(self, filepath: str) -> List[Dict]:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    def match(self) -> List[Dict]:
        """Find matches between Balflex and Heizmann fittings"""
        print(f"\nMatching {len(self.balflex_fittings)} Balflex with {len(self.heizmann_fittings)} Heizmann...")
        print("=" * 70)
        
        # Exclude non-fitting categories
        excluded = [
            'Staub- & Gewindeschutz', 'Hydraulik-Dichtungen', 'Sortimente',
            'Rohrtechnik', 'Messtechnik', 'Leitungszubehör'
        ]
        
        valid_heizmann = [
            f for f in self.heizmann_fittings
            if f.get('category') not in excluded
        ]
        
        print(f"Valid Heizmann: {len(valid_heizmann)} (excluded {len(self.heizmann_fittings) - len(valid_heizmann)})")
        
        # Match
        for bal in self.balflex_fittings:
            best_match = None
            best_score = 0
            
            for heiz in valid_heizmann:
                score = self._calculate_score(bal, heiz)
                
                # Require minimum 55% score and thread compatibility
                if score > best_score and score >= 55:
                    # Check thread compatibility
                    if self._threads_compatible(bal, heiz):
                        best_score = score
                        best_match = heiz
            
            if best_match:
                self.matches.append({
                    'balflex_reference': bal.get('reference', ''),
                    'balflex_category': bal.get('category', ''),
                    'balflex_dash_size': bal.get('dash_size', ''),
                    'balflex_thread': bal.get('thread_type', ''),
                    'balflex_hose_mm': bal.get('hose_size_mm', ''),
                    
                    'heizmann_article': best_match.get('article_number', ''),
                    'heizmann_category': best_match.get('category', ''),
                    'heizmann_thread': best_match.get('thread_type', ''),
                    'heizmann_size': best_match.get('size', ''),
                    'heizmann_dn': best_match.get('DN', ''),
                    'heizmann_pressure': best_match.get('pressure', ''),
                    
                    'match_score': best_score,
                    'match_reason': self._get_match_reason(bal, best_match)
                })
        
        print(f"\n✅ Found {len(self.matches)} matches!")
        return self.matches
    
    def _calculate_score(self, bal: Dict, heiz: Dict) -> int:
        """Calculate match score"""
        score = 0
        
        # Thread type (40 points)
        bal_thread = bal.get('thread_type', '').upper()
        heiz_thread = heiz.get('thread_type', '').upper()
        
        if bal_thread and heiz_thread:
            if bal_thread == heiz_thread:
                score += 40
            elif 'BSP' in bal_thread and 'BSP' in heiz_thread:
                score += 40
            elif 'ORFS' in bal_thread and 'ORFS' in heiz_thread:
                score += 40
            elif 'JIC' in bal_thread and 'JIC' in heiz_thread:
                score += 40
            elif 'METRIC' in bal_thread and 'METRIC' in heiz_thread:
                score += 40
            elif 'NPT' in bal_thread and 'NPT' in heiz_thread:
                score += 40
        
        # Dash size match (30 points)
        bal_dash = bal.get('dash_size', '').replace('-', '')
        heiz_size = heiz.get('size', '')
        
        if bal_dash and heiz_size:
            if bal_dash in heiz_size:
                score += 30
            elif self._size_compatible(bal_dash, heiz_size):
                score += 20
        
        # Category similarity (15 points)
        bal_cat = bal.get('category', '').upper()
        heiz_cat = heiz.get('category', '').upper()
        
        if bal_cat == heiz_cat:
            score += 15
        elif any(word in heiz_cat for word in ['PRESS', 'ADAPTER', 'FLANGE'] if word in bal_cat):
            score += 10
        
        # Hose size (15 points)
        bal_hose = bal.get('hose_size_mm', '')
        heiz_dn = heiz.get('DN', '')
        
        if bal_hose and heiz_dn:
            try:
                bal_mm = float(bal_hose)
                dn_match = re.search(r'\d+', heiz_dn)
                if dn_match:
                    heiz_mm = float(dn_match.group())
                    if abs(bal_mm - heiz_mm) <= 3:
                        score += 15
                    elif abs(bal_mm - heiz_mm) <= 6:
                        score += 10
            except:
                pass
        
        return score
    
    def _threads_compatible(self, bal: Dict, heiz: Dict) -> bool:
        """Check if thread types are compatible"""
        bal_thread = bal.get('thread_type', '').upper()
        heiz_thread = heiz.get('thread_type', '').upper()
        
        if not bal_thread or not heiz_thread:
            return False
        
        # Exact match
        if bal_thread == heiz_thread:
            return True
        
        # Compatible thread families
        if 'BSP' in bal_thread and 'BSP' in heiz_thread:
            return True
        if 'ORFS' in bal_thread and ('ORFS' in heiz_thread or 'UNF' in heiz_thread):
            return True
        if 'JIC' in bal_thread and 'JIC' in heiz_thread:
            return True
        if 'METRIC' in bal_thread and 'METRIC' in heiz_thread:
            return True
        if 'SCHNEIDRING' in bal_thread and 'SCHNEIDRING' in heiz_thread:
            return True
        if 'NPT' in bal_thread and 'NPT' in heiz_thread:
            return True
        
        return False
    
    def _size_compatible(self, dash: str, size: str) -> bool:
        """Check if sizes are compatible"""
        try:
            dash_num = int(dash)
            if str(dash_num) in size:
                return True
        except:
            pass
        return False
    
    def _get_match_reason(self, bal: Dict, heiz: Dict) -> str:
        """Generate match reason"""
        bal_thread = bal.get('thread_type', 'N/A')
        heiz_thread = heiz.get('thread_type', 'N/A')
        bal_dash = bal.get('dash_size', 'N/A')
        heiz_size = heiz.get('size', 'N/A')
        
        return f"Thread: {bal_thread} ↔ {heiz_thread} | Size: {bal_dash} ↔ {heiz_size}"
    
    def save_excel(self, filename: str):
        """Save matches to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Fittings Matches"
        
        # Headers
        headers = [
            'Match Score', 'Match Reason',
            'Balflex Reference', 'Balflex Category', 'Balflex Thread', 'Balflex Dash', 'Balflex Hose (mm)',
            'Heizmann Article', 'Heizmann Category', 'Heizmann Thread', 'Heizmann Size', 'Heizmann DN', 'Heizmann Pressure'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_idx, match in enumerate(self.matches, 2):
            ws.cell(row_idx, 1, match['match_score'])
            ws.cell(row_idx, 2, match['match_reason'])
            ws.cell(row_idx, 3, match['balflex_reference'])
            ws.cell(row_idx, 4, match['balflex_category'])
            ws.cell(row_idx, 5, match['balflex_thread'])
            ws.cell(row_idx, 6, match['balflex_dash_size'])
            ws.cell(row_idx, 7, match['balflex_hose_mm'])
            ws.cell(row_idx, 8, match['heizmann_article'])
            ws.cell(row_idx, 9, match['heizmann_category'])
            ws.cell(row_idx, 10, match['heizmann_thread'])
            ws.cell(row_idx, 11, match['heizmann_size'])
            ws.cell(row_idx, 12, match['heizmann_dn'])
            ws.cell(row_idx, 13, match['heizmann_pressure'])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws.column_dimensions[col].width = 18
        
        wb.save(filename)
        print(f"✅ Saved Excel: {filename}")
    
    def save_json(self, filename: str):
        """Save matches to JSON"""
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(self.matches, f, ensure_ascii=False, indent=2)
        print(f"✅ Saved JSON: {filename}")


if __name__ == "__main__":
    matcher = ImprovedFittingsMatcher(
        'data/balflex_fittings.json',
        'data/heizmann_fittings_merged.json'
    )
    
    matches = matcher.match()
    
    # Analyze
    print(f"\n📊 MATCH ANALYSIS:")
    avg_score = sum(m['match_score'] for m in matches) / len(matches) if matches else 0
    print(f"Average score: {avg_score:.1f}%")
    
    excellent = sum(1 for m in matches if m['match_score'] >= 80)
    good = sum(1 for m in matches if 60 <= m['match_score'] < 80)
    fair = sum(1 for m in matches if 45 <= m['match_score'] < 60)
    
    print(f"Excellent (≥80%): {excellent} ({excellent/len(matches)*100:.1f}%)")
    print(f"Good (60-79%): {good} ({good/len(matches)*100:.1f}%)")
    print(f"Fair (45-59%): {fair} ({fair/len(matches)*100:.1f}%)")
    
    # Save
    matcher.save_excel('fittings_comparison_improved.xlsx')
    matcher.save_json('data/fittings_matches_improved.json')
    
    print(f"\n✅ DONE!")
