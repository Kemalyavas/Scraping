"""
Excel Generator
Creates formatted Excel comparison report from matched products
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelGenerator:
    """Generate formatted Excel report for product comparisons"""
    
    def __init__(self, matches_file: str):
        self.matches_file = matches_file
        self.matches = self._load_matches()
    
    def _load_matches(self):
        """Load matches from JSON file"""
        try:
            with open(self.matches_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            print(f"Error: {self.matches_file} not found!")
            return []
    
    def generate_excel(self, output_file: str):
        """Generate formatted Excel file"""
        if not self.matches:
            print("No matches to export!")
            return
        
        print(f"Generating Excel report with {len(self.matches)} matches...")
        
        # Create DataFrame
        df = self._create_dataframe()
        
        # Sort by match score (descending)
        df = df.sort_values('Match Score %', ascending=False)
        
        # Save to Excel
        df.to_excel(output_file, index=False, sheet_name='Product Comparison')
        
        # Format the Excel file
        self._format_excel(output_file)
        
        print(f"✓ Excel report saved to: {output_file}")
    
    def _create_dataframe(self) -> pd.DataFrame:
        """Create pandas DataFrame from matches"""
        data = []
        
        for match in self.matches:
            row = {
                'Match Score %': match.get('match_score', 0),
                'Match Quality': match.get('match_quality', ''),
                'Match Reasons': match.get('match_reasons', ''),
                
                # Balflex columns
                'Balflex Model': match.get('balflex_model', ''),
                'Balflex Reference': match.get('balflex_reference', ''),
                'Balflex Article #': match.get('balflex_article_number', ''),
                'Balflex Category': match.get('balflex_category', ''),
                
                # Heizmann columns
                'Heizmann Model': match.get('heizmann_model', ''),
                'Heizmann Reference': match.get('heizmann_reference', ''),
                'Heizmann Article #': match.get('heizmann_article_number', ''),
                'Heizmann Category': match.get('heizmann_category', ''),
                
                # Technical specifications
                'DN': match.get('dn', ''),
                'Inch Size': match.get('inch_size', ''),
                'Pressure (MPa)': match.get('working_pressure_mpa', ''),
                'Pressure (bar)': match.get('working_pressure_bar', ''),
                'Pressure (PSI)': match.get('working_pressure_psi', ''),
                'Standard/Norm': match.get('standard', ''),
                'Construction Type': match.get('construction', ''),
                
                # Diameter comparison
                'Balflex ID (mm)': match.get('balflex_inner_diameter_mm', ''),
                'Heizmann ID (mm)': match.get('heizmann_inner_diameter_mm', ''),
                'Balflex OD (mm)': match.get('balflex_outer_diameter_mm', ''),
                'Heizmann OD (mm)': match.get('heizmann_outer_diameter_mm', ''),
            }
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _format_excel(self, file_path: str):
        """Apply formatting to Excel file"""
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        balflex_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6")
        heizmann_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC")
        
        excellent_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE")
        good_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C")
        fair_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE")
        
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Side(border_style="thin", color="000000")
        border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        # Format data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = border
                cell.alignment = left_align if col_idx > 3 else center_align
                
                # Color code by match quality
                if col_idx == 2:  # Match Quality column
                    if cell.value == "Excellent Match":
                        cell.fill = excellent_fill
                    elif cell.value == "Good Match":
                        cell.fill = good_fill
                    elif cell.value in ["Fair Match", "Possible Match"]:
                        cell.fill = fair_fill
                
                # Balflex columns background
                if 4 <= col_idx <= 7:
                    if not cell.fill.start_color or cell.fill.start_color.rgb == '00000000':
                        cell.fill = balflex_fill
                
                # Heizmann columns background
                if 8 <= col_idx <= 11:
                    if not cell.fill.start_color or cell.fill.start_color.rgb == '00000000':
                        cell.fill = heizmann_fill
        
        # Auto-adjust column widths
        column_widths = {
            'A': 12,  # Match Score
            'B': 15,  # Match Quality
            'C': 40,  # Match Reasons
            'D': 25,  # Balflex Model
            'E': 15,  # Balflex Reference
            'F': 18,  # Balflex Article #
            'G': 20,  # Balflex Category
            'H': 25,  # Heizmann Model
            'I': 15,  # Heizmann Reference
            'J': 18,  # Heizmann Article #
            'K': 20,  # Heizmann Category
            'L': 10,  # DN
            'M': 10,  # Inch Size
            'N': 12,  # Pressure MPa
            'O': 12,  # Pressure bar
            'P': 12,  # Pressure PSI
            'Q': 25,  # Standard
            'R': 18,  # Construction
            'S': 12,  # Balflex ID
            'T': 12,  # Heizmann ID
            'U': 12,  # Balflex OD
            'V': 12,  # Heizmann OD
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze header row and first 3 columns
        ws.freeze_panes = 'D2'
        
        # Add summary sheet
        self._add_summary_sheet(wb)
        
        # Save
        wb.save(file_path)
    
    def _add_summary_sheet(self, wb):
        """Add a summary sheet with statistics"""
        ws_summary = wb.create_sheet('Summary', 0)
        
        # Calculate statistics
        total_matches = len(self.matches)
        quality_counts = {}
        
        for match in self.matches:
            quality = match.get('match_quality', 'Unknown')
            quality_counts[quality] = quality_counts.get(quality, 0) + 1
        
        # Category counts
        balflex_categories = {}
        heizmann_categories = {}
        
        for match in self.matches:
            b_cat = match.get('balflex_category', 'Unknown')
            h_cat = match.get('heizmann_category', 'Unknown')
            balflex_categories[b_cat] = balflex_categories.get(b_cat, 0) + 1
            heizmann_categories[h_cat] = heizmann_categories.get(h_cat, 0) + 1
        
        # Write summary
        ws_summary['A1'] = "Hydraulic Hose Product Comparison Summary"
        ws_summary['A1'].font = Font(bold=True, size=14, color="366092")
        
        row = 3
        ws_summary[f'A{row}'] = "Total Matches Found:"
        ws_summary[f'B{row}'] = total_matches
        ws_summary[f'B{row}'].font = Font(bold=True)
        
        row += 2
        ws_summary[f'A{row}'] = "Match Quality Breakdown:"
        ws_summary[f'A{row}'].font = Font(bold=True, underline='single')
        
        row += 1
        for quality in ["Excellent Match", "Good Match", "Fair Match", "Possible Match"]:
            count = quality_counts.get(quality, 0)
            ws_summary[f'A{row}'] = quality
            ws_summary[f'B{row}'] = count
            ws_summary[f'C{row}'] = f"{count/total_matches*100:.1f}%" if total_matches > 0 else "0%"
            row += 1
        
        row += 2
        ws_summary[f'A{row}'] = "Balflex Categories:"
        ws_summary[f'A{row}'].font = Font(bold=True, underline='single')
        row += 1
        for cat, count in sorted(balflex_categories.items()):
            ws_summary[f'A{row}'] = cat
            ws_summary[f'B{row}'] = count
            row += 1
        
        row += 2
        ws_summary[f'A{row}'] = "Heizmann Categories:"
        ws_summary[f'A{row}'].font = Font(bold=True, underline='single')
        row += 1
        for cat, count in sorted(heizmann_categories.items()):
            ws_summary[f'A{row}'] = cat
            ws_summary[f'B{row}'] = count
            row += 1
        
        # Format columns
        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 12
        ws_summary.column_dimensions['C'].width = 12


def main():
    """Test the Excel generator"""
    data_dir = Path(__file__).parent.parent / 'data'
    output_dir = Path(__file__).parent.parent / 'output'
    
    matches_file = data_dir / 'product_matches.json'
    output_file = output_dir / 'product_comparison.xlsx'
    
    if not matches_file.exists():
        print(f"Error: {matches_file} not found!")
        print("Run product_matcher.py first")
        return
    
    generator = ExcelGenerator(str(matches_file))
    generator.generate_excel(str(output_file))
    
    print(f"\n✓ Excel report generated successfully!")
    print(f"  Location: {output_file}")


if __name__ == '__main__':
    main()
